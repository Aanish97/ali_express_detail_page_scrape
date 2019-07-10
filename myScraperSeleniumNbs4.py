#loading the url-html page in html variable
from selenium import webdriver
from bs4 import BeautifulSoup
import csv

url = input('Enter the URL :')

browser = webdriver.Chrome("C:\\Users\Aanish\Downloads\chromedriver_win32\chromedriver.exe")
browser.get(url)

import time
time.sleep(10)
html = browser.page_source

#	print(html)

#loading the html data into the soup variable and the scrapedData.txt
soup = BeautifulSoup(html,'html.parser')

#html_file = open("scrapedData","w", encoding="utf-8")
#html_file.write(html)
#html_file.close()

#Product page Details
productTitle = soup.find("div",{"class":"product-title"}).text.strip()
price = soup.find("div",{"class":"product-price-current"}).text.strip()
overviewRating = soup.find("span",{"class":"overview-rating-average"}).text
noOfReviews = soup.find("span",{"class":"product-reviewer-reviews black-link"}).text
#noOfOrders = soup.find("span",{"class":"product-reviewer-sold"}).text
noInStock = soup.find("div",{"class":"product-quantity-tip"}).text.strip()
shippingType = soup.find("span",{"class":"product-shipping-price bold"}).text
shippingInfo = soup.find("span",{"class":"product-shipping-info black-link"}).text
deliveryTime = soup.find("span",{"class":"product-shipping-delivery"}).text

sDic = {'product Name = ' : productTitle,
		'product Price = ' : price,
		'overview-rating = ' : overviewRating,
'no of reviews = ' : noOfReviews,
#'no of orders = ' : noOfOrders,
'no in Stock = ' : noInStock,
'shippingType = ' : shippingType,
"shippingInfo = " : shippingInfo,
"deliveryTime = " : deliveryTime}

print(sDic)

nxtPage = soup.find("div",{"class":"may-like-item"})
nxtPage = nxtPage.find('a', href=True)
url = 'https:' + nxtPage['href']
print(url)

import pandas
import os
from openpyxl import load_workbook
import xlsxwriter

if __name__ == '__main__':
	get_next_dict = iter([sDic]*1)
	headers = sDic.keys()

	# create csv file if it does not exist
	if not os.path.isfile('test.csv'):
		with open('test.csv', 'w')as csv_file:
			csv_file.writelines(', '.join(headers))

	# create excel file if it does not exist
	if not os.path.isfile('test.xlsx'):
		book = xlsxwriter.Workbook('test.xlsx')
		sheet = book.add_worksheet("TestSheet")
		for (idx, header) in enumerate(headers):
			sheet.write(0, idx, header)
		book.close()

	# open the files and start the loop
	with open('test.csv', 'a+') as csv_file:
		book = load_workbook('test.xlsx')
		sheet = book.get_sheet_by_name('TestSheet')

		# loop through all dictionaries
		for d in get_next_dict:
			values = [d[key] for key in headers]
			csv_string = '\n'+', '.join(values)
			# write to csv file
			csv_file.write(csv_string)
			# write to excel file
			sheet.append(values)
		book.save(filename='test.xlsx')